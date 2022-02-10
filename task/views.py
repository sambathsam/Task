from django.shortcuts import render,HttpResponse
import openpyxl,json
from geopy.geocoders import Nominatim
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from support import supply
import datetime,re
con = supply()

db,cursor = con.mssql(1, 'faretrack') 
dcursor = db.cursor(as_dict = True)  
offdb, offcursor = con.psql(4, 'flights')
inputtable = 'flights.dbo.input_3v'

def inputget():
	status = 0
	Websitecode = 42
	startid = 2;
	endid = 5
	selectq = f"select top 1 id, depart, arrive, triptype, departdate, returndate, adults, url, websitecode from {inputtable} with (nolock) where status = {status} and websitecode={Websitecode} and id between {startid} and {endid} order by id"
	cursor.execute(selectq)
	resultset = cursor.fetchall() 
	for messages in resultset:
		start_time = datetime.datetime.now()
		refid          = messages[0]
		flyfrom        = messages[1]
		flyto          = messages[2]
		routetype      = messages[3]
		departdate     = messages[4]
		returndate     = messages[5]
		adults         = messages[6]
		url            = messages[7]
		source         = messages[8] 
		urlform   = 'https://www.britishairways.com/travel/book/public/en_in/flightList?onds='+flyfrom+'-'+flyto+'_'+str(departdate)+'&ad='+str(adults)+'&yad=0&ch=0&inf=0&cabin=M&flex=LOWEST'
		print("URLL:",urlform)
	return refid,urlform

def getlatlng(addr):
    geolocator = Nominatim(user_agent="sam@gmail.com")
    location   = geolocator.geocode(addr)
    if location:
        return location.latitude, location.longitude
    else:
        return 0,0
    
def home(request):
    try:
        if "GET" == request.method:
            return render(request, 'home.html')
        else:
            excel_file = request.FILES["excel_file"]
            response   = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{excel_file}"'
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb["Sheet1"]
            for row in worksheet.iter_rows():
                for cell in row:
                    xcel_val = getlatlng(str(cell.value))
                    for i,value_ in enumerate(xcel_val,1):
                        worksheet.cell(row=cell.row, column=cell.column+i, value=value_)
            wb.save(response)
            return response
    except Exception as e:
        print("Error:-",e)
        return HttpResponse(f"Error:-{e}")
        
@csrf_exempt        
def get_url(request):
	if "GET" != request.method:
        #return render(request, 'home.html')
		#else:
		print("here the get")
		id_,url = inputget()
		#url= 'https://www.britishairways.com/travel/book/public/en_in/flightList?onds=BLR-LON_2022-02-11&ad=1&yad=0&ch=0&inf=0&cabin=M&flex=LOWEST'
		data = {'id': id_, 'url': url}
		return JsonResponse(data, safe=False)
	else:
		return HttpResponse("Error in GET-URL")
		
@csrf_exempt     
def postdata(request):
	if request.method =='POST':
		print("here the data comes")
		body_unicode = request.body.decode('utf-8')
		body = json.loads(body_unicode)
		urlins = body['url']
		fdata  = body['data']
		id_up  = body['id']
		print(urlins)
		flyfrom  = re.search(r"flightList\?\w+=(.*?)\-",urlins).group(1)
		departuredate = re.search(r"flightList\?\w+=.*?\-.*?_(.*?)&",urlins).group(1)
		websitecode = 42
		offlinesource   = re.sub("'", "''", str(fdata))
		insertq         = "insert into victor_offline (refid,websitecode,data,flyfrom,departuredate) values ('%s', '%s', '%s','%s','%s')"%(id_up,websitecode, offlinesource,flyfrom,departuredate)
		offcursor.execute(insertq)
		offdb.commit()
		print("data inserted-id-",id_up)
		cursor.execute("update "+inputtable+" set status = 1 where id= %s"%(id_up))
		db.commit()
		print('updated web',websitecode,'And id--',id_up)
		return request

	else:
		return HttpResponse("Error in POST-DATA-")
    
    
    
